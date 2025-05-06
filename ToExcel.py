"""
UNResolutionProcessor: ToExcel.py

    exports processed data to .xlsx (initialize/truncate and overwrite)

"""
import os
import pandas as pd
from openpyxl import load_workbook
from ProcessHelpers import animated_printer

def add_dataframe_to_excel(file_path, sheet_name, df, startrow=None, truncate_sheet=True, **to_excel_kwargs):
    try:
        # check if the file exists
        if os.path.exists(file_path):
            # load existing workbook
            writer = pd.ExcelWriter(file_path, engine='openpyxl')
            writer.book = load_workbook(file_path)

            # if sheet exists, determine start row
            if sheet_name in writer.book.sheetnames:
                sheet = writer.book[sheet_name]
                if startrow is None:
                    startrow = sheet.max_row if sheet.max_row is not None else 0

                # truncate sheet if needed
                if truncate_sheet:
                    idx = writer.book.sheetnames.index(sheet_name)
                    writer.book.remove(writer.book.worksheets[idx])
                    writer.book.create_sheet(sheet_name, idx)
                    startrow = 0
            else:
                startrow = 0
        else:
            # create a new workbook
            writer = pd.ExcelWriter(file_path, engine='openpyxl')
            startrow = 0

        # write dataframe to .xlsx
        df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, **to_excel_kwargs)

        # save the workbook
        writer.close()
        animated_printer.safe_print(f"DataFrame added to {file_path} in sheet '{sheet_name}'.")
    except Exception as e:
        animated_printer.safe_print(f"An error occurred: {e}")

def overwrite_excel_file(file_path, sheet_name, df, **to_excel_kwargs):
    try:
        # write dataframe to .xlsx, overwriting a file if it exists at that filename
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, **to_excel_kwargs)
        animated_printer.safe_print(f"DataFrame written to {file_path} in sheet '{sheet_name}'.")
    except Exception as e:
        animated_printer.safe_print(f"An error occurred: {e}")